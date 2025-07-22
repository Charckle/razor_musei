from openpyxl import Workbook
from openpyxl.styles import Font

from datetime import datetime


from io import BytesIO
from flask import make_response

from app.main_page_module.p_objects.artifact import Artifact


def makeResponse(content, fileName):
    resp = make_response(content.read())
    resp.headers["Content-Disposition"] = f"attachment; filename={fileName}.xlsx"
    resp.headers['Content-Type'] = 'application/x-xlsx'
    
    return resp


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter  # Get the column letter (e.g., A, B, C)
        for cell in column_cells:
            try:
                if cell.value:  # Skip empty cells
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column_letter].width = adjusted_width

def fill_sheet(ws, name, data):
    ws.title = name
    
    bold_font = Font(bold=True)      
    
    ws[f'A1'] = "Ime"
    ws['A1'].font = bold_font  
    ws[f'B1'] = "col_ref_num"
    ws['B1'].font = bold_font
    ws[f'C1'] = "Tip"
    ws['C1'].font = bold_font
    ws[f'D1'] = "Obdobje"
    ws['D1'].font = bold_font
    ws[f'E1'] = "Politicna Entiteta"
    ws['E1'].font = bold_font
    ws[f'F1'] = "Replika"
    ws['F1'].font = bold_font
    ws[f'G1'] = "Cena ob nakupu €"
    ws['G1'].font = bold_font
    ws[f'H1'] = "Cena ob prodaji €"
    ws['H1'].font = bold_font
    ws[f'I1'] = "Leto prejema v kolekcijo"
    ws['I1'].font = bold_font
    ws[f'J1'] = "Leto oddaje iz kolekcije"
    ws['J1'].font = bold_font
    ws[f'K1'] = "Javno vidno"
    ws['K1'].font = bold_font
    ws[f'L1'] = "Trenutna lokacija artifakta"
    ws['L1'].font = bold_font
    

    ws[f'M1'] = "Tip kovanca"
    ws['M1'].font = bold_font
    ws[f'N1'] = "Vladar, pod katerim je bila kovanec izdan"
    ws['N1'].font = bold_font
    ws[f'O1'] = "Mesto kovnice"
    ws['O1'].font = bold_font
    ws[f'P1'] = "Leto/obdobje kovanja"
    ws['P1'].font = bold_font
    ws[f'Q1'] = "Material"
    ws['Q1'].font = bold_font
    ws[f'R1'] = "Teza"
    ws['R1'].font = bold_font
    ws[f'S1'] = "Premer"
    ws['S1'].font = bold_font
    ws[f'T1'] = "Ocena Stanja"
    ws['T1'].font = bold_font
    ws[f'U1'] = "Opis kovanca"
    ws['U1'].font = bold_font
    ws[f'V1'] = "Obverse - Lice"
    ws['V1'].font = bold_font
    ws[f'W1'] = "Reverse - Hrbet"
    ws['W1'].font = bold_font        
    
    
    ws[f'X1'] = "Reference"
    ws['X1'].font = bold_font
    ws[f'Y1'] = "Opombe o izvoru"
    ws['Y1'].font = bold_font
    ws[f'Z1'] = "Lastniki"
    ws['Z1'].font = bold_font
    ws[f'AA1'] = "Opis"
    ws['AA1'].font = bold_font
    ws[f'AB1'] = "Zgodovinski kontekts"
    ws['AB1'].font = bold_font
    
    
    #table
    first_row = 2
    for col_ref_num, artifact in data.items():
        ws[f'A{first_row}'] = artifact['name']
        ws[f'B{first_row}'] = artifact['col_ref_num']
        ws[f'C{first_row}'] = Artifact.types(artifact["type_"])
        ws[f'D{first_row}'] = Artifact.periods(artifact["period"])
        ws[f'E{first_row}'] = artifact['state_entity']
        ws[f'F{first_row}'] = Artifact.yes_no(artifact["replica"])
        ws[f'G{first_row}'] = artifact['buy_price']
        ws[f'H{first_row}'] = artifact['sold_price']
        ws[f'I{first_row}'] = artifact['joined_collection_in_year']
        ws[f'J{first_row}'] = artifact['left_collection_in_year']
        ws[f'K{first_row}'] = Artifact.yes_no(artifact["public"])
        ws[f'L{first_row}'] = artifact['curr_location_of_item']
        ws[f'M{first_row}'] = artifact['coin_type']
        ws[f'N{first_row}'] = artifact['ruler']
        ws[f'O{first_row}'] = artifact['mint_city']
        ws[f'P{first_row}'] = artifact['mint_period']
        ws[f'Q{first_row}'] = artifact['material']
        ws[f'R{first_row}'] = artifact['weight']
        ws[f'S{first_row}'] = artifact['diameter']
        ws[f'T{first_row}'] = Artifact.grades(artifact["grade"])
        ws[f'U{first_row}'] = artifact['coin_description']
        ws[f'V{first_row}'] = artifact['obverse']
        ws[f'W{first_row}'] = artifact['reverse']
        ws[f'X{first_row}'] = artifact['reference']
        ws[f'Y{first_row}'] = artifact['provenance_notes']
        ws[f'Z{first_row}'] = artifact['reverse']
        ws[f'AA{first_row}'] = artifact['description']
        ws[f'AB{first_row}'] = artifact['historical_context']
        
        first_row += 1
    
    # Automatically adjust column widths
    auto_fit_columns(ws)    
    

class ExcelO:
    # ExcelO
    @staticmethod    
    def export_artifacts(artifacts):
        current_year = datetime.now().year
        fileName = f"Zgodovinska_zbirka_artifakti_{current_year }"
        
        # Create a new Workbook
        wb = Workbook()
        
        # Access the active sheet
        ws = wb.active
        artifacts_sorted = {k: v for k, v in artifacts.items() if v["type_"] != "coin"}
        fill_sheet(ws, "Artifacts", artifacts_sorted)
        
        ws_name = "Coins"
        new_ws = wb.create_sheet(title=ws_name)
        artifacts_sorted = {k: v for k, v in artifacts.items() if v["type_"] == "coin"}
        fill_sheet(new_ws, ws_name, artifacts_sorted)
        
        
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        wordBook = makeResponse(output, fileName)
        
        return wordBook

